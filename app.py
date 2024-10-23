import openpyxl, os
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui 

webbrowser.open('https://web.whatsapp.com/')
sleep(30)

if not os.path.exists('clientes.xlsx'): #Se não existir clientes.xlsx na pasta onde tá o script
    print('Arquivo não encontrado no diretório do script!') #Imprime isso

try: # Tenta executar o código abaixo, se der erro pula pro except
    workbook = openpyxl.load_workbook('clientes.xlsx')
    valor = int(input('Digite o min_row: ')) # Pede o min row para recortar o dado a partir da linha que digitarem
    pagina_clientes = workbook['Planilha1']
    # pagina_clientes.iter_cols(): # iterator de cols
    for linha in pagina_clientes.iter_rows(min_row=valor):
        if all((celula.value is None or str(celula.value).strip() == "") for celula in linha):
            continue

        # if not linha:
        #     continue
        # Abaixo salva nas variáveis respectivas as colunas caso o tamanho da variável linha seja maior que X
        nome = linha[0].value if len(linha) > 0 else None
        telefone = linha[1].value if len(linha) > 1 else None
    
        print(f'Nome: {nome}')
        print(f'Telefone: {telefone}')

        mensagem = f'Olá {nome},tudo bem? Acesse o link abaixo, siga nosso Instagram.'

        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'

        webbrowser.open(link_mensagem_whatsapp)
        sleep(10) 
        seta = pyautogui.locateCenterOnScreen('setinhawpp.png')
        sleep(5)
        pyautogui.click(seta[0],seta[1])
        sleep(5)
        pyautogui.hotkey('ctrl','w')
        sleep(5)

        
except Exception as error: # Oi aqui é o except
    print(f'Não foi possível enviar mensagem para: {error}')
    with open('erros.csv','a','newline=', enconding='utf-8')as arquivo:
        arquivo.write(f'{nome},{telefone}')




